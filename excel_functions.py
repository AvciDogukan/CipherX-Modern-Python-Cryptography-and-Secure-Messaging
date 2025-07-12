# -*- coding: utf-8 -*-
"""
Created on Sun Aug 11 20:48:42 2024

@author: Dogukan Avcı
Excel Functions
"""

# excel_functions.py
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

def log_to_excel(message, key):
    filename = "CipherX_encryption_log.xlsx"
    
    try:
        # Mevcut dosya varsa aç
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # Dosya yoksa yeni bir çalışma kitabı oluştur
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Cipher X Encryption Log"

        # Başlıklar eklenmemişse, ekle ve stil ver
        headers = ["Mesaj", "Anahtar"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    sheet = workbook.active

    # Mesaj ve anahtarı ekle
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1, value=message)
    sheet.cell(row=new_row, column=2, value=key)

    for col_num in range(1, 3):
        cell = sheet.cell(row=new_row, column=col_num)
        cell.font = Font(size=12)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Sütun genişliğini ayarla
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20

    # Dosyayı kaydet
    workbook.save(filename)

def open_excel_file():
    filename = "CipherX_encryption_log.xlsx"
    if os.path.exists(filename):
        os.startfile(filename)
    else:
        print("Dosya bulunamadı.")
